from openpyxl import load_workbook, Workbook

import configs

START_LOAD_DATA_COL = 2


def load_data(file_path="input.xlsx"):
    values = []

    try:
        wb = load_workbook(file_path)

        counter = START_LOAD_DATA_COL
        while True:
            value = wb[wb.sheetnames[0]]['A' + str(counter)].value
            if not value:
                break
            values.append(value)
            counter += 1

    except FileNotFoundError:
        print('Отсутствует входной файл')
        exit()

    return values


def save_data(products, with_image=False, file_path="output.xlsx"):
    wb = Workbook()
    not_found_products = []

    sheet = wb.active

    sheet.title = "Товары"
    rows = ['Код', "Название", "Стоимость", "Ссылка", "Вес упаковки", "Длина", "Ширина", "Высота"]
    if with_image:
        rows.append('Изображение')
    sheet.append(rows)

    for product in products:
        if product['name'] == configs.NOT_FOUND_STR:
            not_found_products.append(product)
            continue

        output_list = [
            product['id'],
            product['name'],
            float(product['price'].replace(',', '.')) if configs.NOT_FOUND_STR != product['price'] else product['price'],
            '=HYPERLINK("' + product['url'] + '","' + product['url'] + '")' if product['url'] != configs.NOT_FOUND_STR else product['url'],
            float(product['weight'].replace(',', '.')) if configs.NOT_FOUND_STR != product['weight'] else product['weight']]

        try:
            output_list = output_list + [float(item) for item in product['size']]
        except BaseException:
            output_list.append([configs.NOT_FOUND_STR for _ in range(3)])

        if with_image:
            output_list.append('=HYPERLINK("file://' + product['image'] + '","' + product['image'] + '")' if product['image'] != configs.NOT_FOUND_STR else product['image'])

        try:
            sheet.append(output_list)
        except BaseException:
            not_found_products.append(product)

    if len(not_found_products) > 0:
        not_found_list = wb.create_sheet(title="Не найденные")
        not_found_list.append(['Код'])
        for product in not_found_products:
            not_found_list.append([product['id']])

    wb.save(filename=file_path)